# -*- coding:utf-8 -*-
import json
import sys
from openpyxl import Workbook

#json_file_name = "world_Demo.geo"
json_file_name = sys.argv[1]
print "json_file_name : ",json_file_name[0:-4]
str_list = json_file_name.split("\\")
print str_list[-1]
file_name = str_list[-1]

file_path = json_file_name[0:json_file_name.find(file_name)]
print "file_path : ",file_path


BEGIN_INDEX = 100000

navi_point_dict = {}
road_dict = {}

class navi_point:
	serial_id = 0
	position = []
	connection_id = []
	scale = 0.0 
	def __init__(self):
		self.serial_id = 0
		self.position = []
		self.connection_id = []
		self.scale = float(0.0)

class road_info:
	lod = 0
	scale = []
	indices = []
	positions = []
	def __init__(self):
		self.road_index = int(0)
		self.lod = int(0)
		self.scale = []
		self.indices = []
		self.positions = []

def load_and_parse_json_road():
	f = open(json_file_name, 'r') 
	setting = json.load(f)
	#print setting
	indices_list = setting[13][1][1] 
	#print indices_list
	pos_list = setting[15][1][0][1][7][5]
	#print len(pos_list)
	scale_list = setting[15][1][1][1][7][5][0]
	#print scale_list
	lod_list = setting[15][3][0][1][7][5][0]
	#print lod_list
	road_len_list = setting[17][0][1][5]
	#print road_len_list

	indices_begin = 0
	indices_end = 0
	for road_index in range(len(road_len_list)):
		road = road_info()
		indices_end += road_len_list[road_index]
		road.lod = lod_list[road_index]		

		for indice in range(indices_begin,indices_end):			
			indice_index = indices_list[indice]
			#print indice_index
			road.indices.append(indice_index)
			#print pos_list[indice_index]
			road.positions.append(pos_list[indice_index])
			#print scale_list[indice_index]
			road.scale.append(scale_list[indice_index])
		indices_begin = indices_end
		#print "road_index",road_index
		road_dict[road_index] = road
		#print road_dict[road_index].indices
	conver_to_navipoint_data(road_dict, pos_list, scale_list,len(pos_list))

def conver_to_navipoint_data(road_d, pos_list, scale_list, numPoint):
	for serial_id in range(numPoint):
		#print "serial_id",serial_id
		point = navi_point()
		point.serial_id = serial_id + 1 + BEGIN_INDEX
		#print pos_list
		point.position = pos_list[serial_id]
		point.connection_id = find_connection_id_list(road_d,serial_id)
		point.scale = scale_list[serial_id]
		navi_point_dict[serial_id] = point
		#print navi_point_dict[serial_id].connection_id
	save_to_excel(navi_point_dict)



def find_connection_id_list(road_d,index):
	connect_point_list = []
	for road_id in range(len(road_d)):
		if index in road_d[road_id].indices:
			connect_index = road_d[road_id].indices.index(index)
			#print road_d[road_id].indices.index(index)
			if connect_index > 0:
				connect_point_list.append(road_d[road_id].indices[connect_index - 1] + 1 + BEGIN_INDEX)
			if connect_index < len(road_d[road_id].indices)-1:
				connect_point_list.append(road_d[road_id].indices[connect_index + 1] + 1 + BEGIN_INDEX)
	#print connect_point_list
	return connect_point_list

def save_to_excel(navi_point_dict):
	workbook = Workbook()
	booksheet = workbook.active
	booksheet.title = u"自动路点"
	# first row
	row0 = [u"路点流水号id",u"路点类型编号",u"备注",u"路点坐标",u"关联路点id",u"是否复活点",u"关联轻功路点",u"轻功路点动作",u"轻功路点距离", u"路点宽度"]
	for i in range(0,len(row0)):
		booksheet.cell(1, i + 1, row0[i]) 
	
	row1 = ["serial_id","No","notes","position","connection_id","if_resurrection_point","flying_connection_point","flying_anim","flying_range", "point_scale"]
	for i in range(0,len(row1)):
		booksheet.cell(2, i + 1, row1[i])

	row2 = ["int","int","str","list(float, 3)","list(int, N)","int","list(int, N)","str","int","float"]
	for i in range(0,len(row2)):
		booksheet.cell(3, i + 1, row2[i])

	for key in navi_point_dict:
		booksheet.cell(int(key) + 4, 1).value = str(key + 1)
		booksheet.cell(int(key) + 4, 2).value =  0

		pos = navi_point_dict[key].position
		point_str = str(pos[0]) + ";" + str(pos[1]) + ";" + str(pos[2])
		#print point_str
		booksheet.cell(int(key) + 4, 4).value = point_str

		connection_array = navi_point_dict[key].connection_id
		connection_id_str = str(connection_array)
		index = len(connection_id_str)
		connection_id_str = connection_id_str[1:index-1]
		connection_id_str = connection_id_str.replace(", ",";")
		#print connection_id_str		
		booksheet.cell(int(key) + 4, 5).value = connection_id_str

		booksheet.cell(int(key) + 4, 10).value = navi_point_dict[key].scale
		
	#workbook.save(u"../../../design/1System/10副本/副本数据/1002_竹林/1002_竹林_ins_navipoint_auto.xlsx")
	workbook.save(file_path + "auto_navipoint_world_" + file_name + ".xlsx")

load_and_parse_json_road()
